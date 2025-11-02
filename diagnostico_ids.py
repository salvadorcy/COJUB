#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Diagnóstico: Comparación de IDs Excel vs Base de Datos

Este script muestra:
1. Los primeros 10 IDs del Excel
2. Los primeros 10 IDs de la BD
3. Qué IDs están en Excel pero no en BD
4. Qué IDs están en BD pero no en Excel

USO:
    python diagnostico_ids.py
"""

import os
import sys
from dotenv import load_dotenv
import pyodbc
import openpyxl

def main():
    print("\n" + "="*70)
    print("🔍 DIAGNÓSTICO DE IDs - Excel vs Base de Datos")
    print("="*70 + "\n")
    
    # 1. Leer IDs del Excel
    print("📂 Leyendo IDs del Excel...")
    try:
        wb = openpyxl.load_workbook('Socis-2025.xlsx')
        ws = wb['Hoja1']
        
        ids_excel = []
        for row in range(2, min(ws.max_row + 1, 20)):  # Primeros 20
            codi = ws.cell(row, 1).value
            if codi:
                codi_str = str(codi).strip()
                ids_excel.append(codi_str)
        
        print(f"✅ Total de IDs en Excel (primeros 18): {len(ids_excel)}")
        print("\n📋 Primeros 10 IDs del Excel:")
        for i, id_excel in enumerate(ids_excel[:10], 1):
            print(f"   {i:2d}. '{id_excel}' (tipo: {type(id_excel).__name__}, longitud: {len(id_excel)})")
        
    except Exception as e:
        print(f"❌ Error al leer Excel: {e}")
        sys.exit(1)
    
    # 2. Leer IDs de la BD
    print("\n💾 Leyendo IDs de la Base de Datos...")
    try:
        # Cargar .env
        load_dotenv('models/.env')
        
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={os.getenv('SQL_SERVER')};"
            f"DATABASE={os.getenv('SQL_DATABASE')};"
            f"UID={os.getenv('SQL_USER')};"
            f"PWD={os.getenv('SQL_PASSWORD')};"
        )
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        # Obtener primeros 10 IDs de la BD
        cursor.execute("SELECT TOP 20 FAMID FROM scazorla_sa.G_Socis ORDER BY FAMID")
        ids_bd_sample = [row[0] for row in cursor.fetchall()]
        
        print(f"✅ Total de IDs en BD (primeros 20): {len(ids_bd_sample)}")
        print("\n📋 Primeros 10 IDs de la BD:")
        for i, id_bd in enumerate(ids_bd_sample[:10], 1):
            print(f"   {i:2d}. '{id_bd}' (tipo: {type(id_bd).__name__}, longitud: {len(id_bd) if id_bd else 0})")
        
        # Obtener todos los IDs de la BD para comparación
        cursor.execute("SELECT FAMID FROM scazorla_sa.G_Socis")
        todos_ids_bd = set(row[0].strip() if row[0] else "" for row in cursor.fetchall())
        
        conn.close()
        
    except Exception as e:
        print(f"❌ Error al leer BD: {e}")
        sys.exit(1)
    
    # 3. Comparar IDs
    print("\n" + "="*70)
    print("🔍 COMPARACIÓN DE IDs")
    print("="*70)
    
    ids_excel_set = set(ids_excel[:18])
    ids_bd_sample_set = set(id_bd.strip() if id_bd else "" for id_bd in ids_bd_sample[:18])
    
    # IDs en Excel pero no en BD
    solo_en_excel = ids_excel_set - todos_ids_bd
    print(f"\n❌ IDs en Excel pero NO en BD: {len(solo_en_excel)}")
    if solo_en_excel:
        for id_ex in list(solo_en_excel)[:5]:
            print(f"   - '{id_ex}'")
        if len(solo_en_excel) > 5:
            print(f"   ... y {len(solo_en_excel) - 5} más")
    
    # IDs en BD pero no en Excel (estos serían marcados como baja)
    solo_en_bd = todos_ids_bd - ids_excel_set
    print(f"\n⚠️  IDs en BD pero NO en Excel (se marcarían como baja): {len(solo_en_bd)}")
    if solo_en_bd:
        for id_bd in list(solo_en_bd)[:10]:
            print(f"   - '{id_bd}'")
        if len(solo_en_bd) > 10:
            print(f"   ... y {len(solo_en_bd) - 10} más")
    
    # IDs que coinciden
    coinciden = ids_excel_set & todos_ids_bd
    print(f"\n✅ IDs que coinciden: {len(coinciden)}")
    if coinciden:
        for id_coin in list(coinciden)[:5]:
            print(f"   - '{id_coin}'")
    
    # 4. Análisis de formato
    print("\n" + "="*70)
    print("🔍 ANÁLISIS DE FORMATO")
    print("="*70)
    
    # Verificar si hay espacios o diferencias
    print("\n🔎 Verificando diferencias de formato...")
    
    # Ejemplo con el primer ID de cada uno
    if ids_excel and ids_bd_sample:
        id_ex = ids_excel[0]
        id_bd = ids_bd_sample[0]
        
        print(f"\nPrimer ID Excel: '{id_ex}'")
        print(f"  - Bytes: {id_ex.encode('utf-8')}")
        print(f"  - Tiene espacios al inicio: {id_ex != id_ex.lstrip()}")
        print(f"  - Tiene espacios al final: {id_ex != id_ex.rstrip()}")
        
        print(f"\nPrimer ID BD: '{id_bd}'")
        print(f"  - Bytes: {id_bd.encode('utf-8')}")
        print(f"  - Tiene espacios al inicio: {id_bd != id_bd.lstrip()}")
        print(f"  - Tiene espacios al final: {id_bd != id_bd.rstrip()}")
    
    # Resumen final
    print("\n" + "="*70)
    print("📊 RESUMEN")
    print("="*70)
    print(f"Total IDs en Excel (muestra): {len(ids_excel_set)}")
    print(f"Total IDs en BD: {len(todos_ids_bd)}")
    print(f"IDs que coinciden: {len(coinciden)}")
    print(f"IDs solo en Excel: {len(solo_en_excel)}")
    print(f"IDs solo en BD (se marcarían BAJA): {len(solo_en_bd)}")
    print("="*70 + "\n")
    
    if len(solo_en_bd) == len(todos_ids_bd):
        print("⚠️  PROBLEMA DETECTADO:")
        print("   NINGÚN ID del Excel coincide con los de la BD")
        print("   Esto causaría que TODOS los socios se marquen como baja")
        print("\n💡 Posibles causas:")
        print("   1. Los IDs en Excel tienen formato diferente (ej: '001' vs '1')")
        print("   2. Los IDs en BD tienen espacios extra")
        print("   3. El Excel no es el correcto")
        print("   4. La columna de IDs en el Excel no es la correcta")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n\nPresiona ENTER para salir...")