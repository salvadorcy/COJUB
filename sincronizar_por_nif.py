#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Sincronización de Socios por NIF/DNI
===============================================
Versión mejorada que identifica socios por NIF en lugar de por código.

Este script:
1. Lee los socios activos del archivo Excel
2. Usa el NIF como identificador único (no el código FAMID)
3. Actualiza o inserta los socios en la base de datos
4. Marca como baja (bBaixa=1) los socios que NO aparecen en el Excel
5. Crea backup automático antes de sincronizar
6. Genera un reporte detallado de los cambios

Autor: Sistema de Gestión COJUB
Fecha: 2024
"""

import os
import sys
import csv
from datetime import datetime, timedelta
from dotenv import load_dotenv
import pyodbc
import openpyxl

class SincronizadorSociosPorNIF:
    def __init__(self, excel_path, env_path='models/.env'):
        """
        Inicializa el sincronizador.
        
        Args:
            excel_path (str): Ruta al archivo Excel con los socios activos
            env_path (str): Ruta al archivo .env con credenciales de BD
        """
        self.excel_path = excel_path
        self.conn = None
        self.stats = {
            'nuevos': 0,
            'actualizados': 0,
            'marcados_baja': 0,
            'sin_cambios': 0,
            'sin_nif': 0,
            'errores': 0,
            'total_excel': 0,
            'total_bd_antes': 0,
            'total_bd_despues': 0,
            'nifs_duplicados': 0
        }
        
        # Cargar variables de entorno desde la ruta especificada
        if os.path.exists(env_path):
            load_dotenv(env_path)
        
        self.conectar_bd()
    
    def conectar_bd(self):
        """Establece conexión con SQL Server."""
        try:
            conn_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={os.getenv('SQL_SERVER')};"
                f"DATABASE={os.getenv('SQL_DATABASE')};"
                f"UID={os.getenv('SQL_USER')};"
                f"PWD={os.getenv('SQL_PASSWORD')};"
            )
            self.conn = pyodbc.connect(conn_str)
            print("✅ Conexión a la base de datos establecida correctamente")
        except pyodbc.Error as ex:
            print(f"❌ Error al conectar a la base de datos: {ex}")
            sys.exit(1)
    
    def limpiar_nif(self, nif):
        """
        Limpia y normaliza el NIF/DNI para comparación.
        
        Args:
            nif: NIF a limpiar
            
        Returns:
            str: NIF limpio en mayúsculas sin espacios ni guiones
        """
        if nif is None or nif == '':
            return None
        # Convertir a string, quitar espacios, guiones y poner en mayúsculas
        nif_clean = str(nif).strip().replace(' ', '').replace('-', '').upper()
        return nif_clean if nif_clean else None
    
    def crear_backup(self):
        """Crea un backup de todos los socios antes de sincronizar."""
        print("\n" + "="*70)
        print("💾 CREANDO BACKUP DE SEGURIDAD")
        print("="*70)
        
        # Crear carpeta de backups si no existe
        backup_dir = "backups"
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
            print(f"📁 Carpeta '{backup_dir}' creada")
        
        # Nombre del archivo de backup con fecha y hora
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(backup_dir, f"backup_socios_{timestamp}.csv")
        
        try:
            cursor = self.conn.cursor()
            cursor.execute("SELECT * FROM scazorla_sa.G_Socis")
            
            # Obtener nombres de columnas
            columns = [column[0] for column in cursor.description]
            
            # Escribir al CSV
            with open(backup_file, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(columns)  # Encabezados
                
                count = 0
                for row in cursor.fetchall():
                    writer.writerow(row)
                    count += 1
            
            print(f"✅ Backup creado: {backup_file}")
            print(f"📊 Total de socios respaldados: {count}")
            print("="*70)
            
            return backup_file
            
        except Exception as e:
            print(f"❌ Error al crear backup: {e}")
            return None
    
    def leer_excel(self):
        """
        Lee el archivo Excel y retorna una lista de diccionarios con los socios.
        Ahora usa el NIF como identificador principal.
        
        Returns:
            list: Lista de diccionarios con datos de socios
        """
        print(f"\n📂 Leyendo archivo Excel: {self.excel_path}")
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb['Hoja1']
            
            socios_excel = []
            nifs_vistos = set()
            
            # Leer desde la fila 2 (saltando encabezados)
            for row_num in range(2, ws.max_row + 1):
                # Obtener valores
                codi = ws.cell(row_num, 1).value  # Codi
                nombre = ws.cell(row_num, 2).value  # Nombre
                nif = ws.cell(row_num, 3).value  # NIF
                direccion = ws.cell(row_num, 4).value  # Dirección
                cp = ws.cell(row_num, 5).value  # CP
                poblacion = ws.cell(row_num, 7).value  # Población
                telefono = ws.cell(row_num, 9).value  # Teléfono
                movil = ws.cell(row_num, 10).value  # Móvil
                email = ws.cell(row_num, 11).value  # Email
                forma_pago = ws.cell(row_num, 12).value  # Forma de pagament
                iban = ws.cell(row_num, 13).value  # IBAN
                bic = ws.cell(row_num, 14).value  # BIC
                fecha_alta = ws.cell(row_num, 15).value  # Fecha Alta
                
                # Limpiar y validar NIF
                nif_limpio = self.limpiar_nif(nif)
                
                if not nif_limpio:
                    self.stats['sin_nif'] += 1
                    print(f"  ⚠️ Fila {row_num}: NIF vacío - Socio '{nombre}' - OMITIDO")
                    continue
                
                # Detectar NIFs duplicados en Excel
                if nif_limpio in nifs_vistos:
                    self.stats['nifs_duplicados'] += 1
                    print(f"  ⚠️ Fila {row_num}: NIF duplicado: {nif_limpio} - Se usará última aparición")
                
                nifs_vistos.add(nif_limpio)
                
                # Convertir Codi a string
                codi = str(codi).strip() if codi else ""
                
                # Determinar si el pago está domiciliado
                pago_domiciliado = False
                if forma_pago and isinstance(forma_pago, str):
                    pago_domiciliado = 'domiciliat' in forma_pago.lower() or '3' in forma_pago
                
                # Convertir fecha de Excel a datetime si es numérico
                fecha_alta_dt = None
                if fecha_alta:
                    if isinstance(fecha_alta, (int, float)):
                        try:
                            fecha_alta_dt = datetime(1899, 12, 30) + timedelta(days=int(fecha_alta))
                        except:
                            fecha_alta_dt = None
                    elif isinstance(fecha_alta, datetime):
                        fecha_alta_dt = fecha_alta
                
                # Convertir CP a string
                cp_str = str(cp) if cp else ""
                if cp_str and '.' in cp_str:
                    cp_str = cp_str.split('.')[0]  # Quitar decimales si los hay
                
                socio = {
                    'FAMID': codi,
                    'NIF_LIMPIO': nif_limpio,  # NIF normalizado para búsqueda
                    'FAMNom': str(nombre).strip() if nombre else "",
                    'FAMNIF': nif_limpio,  # Usar NIF limpio
                    'FAMAdressa': str(direccion).strip() if direccion else "",
                    'FAMPoblacio': str(poblacion).strip() if poblacion else "",
                    'FAMCodPos': cp_str.strip(),
                    'FAMTelefon': str(telefono).strip() if telefono else "",
                    'FAMMobil': str(movil).strip() if movil else "",
                    'FAMEmail': str(email).strip() if email else "",
                    'FAMIBAN': str(iban).strip() if iban else "",
                    'FAMBIC': str(bic).strip() if bic else "",
                    'FAMDataAlta': fecha_alta_dt,
                    'FAMbPagamentDomiciliat': pago_domiciliado,
                    'bBaixa': False  # Los del Excel están activos
                }
                
                socios_excel.append(socio)
            
            self.stats['total_excel'] = len(socios_excel)
            print(f"✅ Se leyeron {len(socios_excel)} socios válidos del Excel")
            if self.stats['sin_nif'] > 0:
                print(f"⚠️  {self.stats['sin_nif']} socios omitidos por NIF vacío")
            if self.stats['nifs_duplicados'] > 0:
                print(f"⚠️  {self.stats['nifs_duplicados']} NIFs duplicados detectados")
            
            return socios_excel
            
        except Exception as e:
            print(f"❌ Error al leer el Excel: {e}")
            sys.exit(1)
    
    def obtener_socios_bd_por_nif(self):
        """
        Obtiene todos los socios de la BD indexados por NIF.
        
        Returns:
            dict: Diccionario {NIF_LIMPIO: {datos del socio}}
        """
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT FAMID, FAMNIF, FAMNom, bBaixa 
            FROM scazorla_sa.G_Socis
        """)
        
        socios_bd = {}
        total_con_nif = 0
        total_sin_nif = 0
        
        for row in cursor.fetchall():
            famid, famnif, famnom, bbaixa = row
            nif_limpio = self.limpiar_nif(famnif)
            
            if nif_limpio:
                socios_bd[nif_limpio] = {
                    'FAMID': famid,
                    'FAMNIF_ORIGINAL': famnif,
                    'FAMNom': famnom,
                    'bBaixa': bbaixa
                }
                total_con_nif += 1
            else:
                total_sin_nif += 1
        
        self.stats['total_bd_antes'] = total_con_nif + total_sin_nif
        print(f"📊 Total de socios en BD: {self.stats['total_bd_antes']}")
        print(f"   • Con NIF válido: {total_con_nif}")
        if total_sin_nif > 0:
            print(f"   • Sin NIF: {total_sin_nif} (no se procesarán)")
        
        return socios_bd
    
    def insertar_socio(self, socio):
        """Inserta un nuevo socio en la base de datos."""
        try:
            cursor = self.conn.cursor()
            
            # Obtener el siguiente FAMID disponible
            cursor.execute("SELECT ISNULL(MAX(FAMID), 0) + 1 FROM scazorla_sa.G_Socis")
            nuevo_famid = cursor.fetchone()[0]
            
            query = """
            INSERT INTO scazorla_sa.G_Socis (
                FAMID, FAMNom, FAMAdressa, FAMPoblacio, FAMCodPos, FAMTelefon,
                FAMMobil, FAMEmail, FAMDataAlta, FAMCCC, FAMIBAN, FAMBIC,
                FAMNSocis, bBaixa, FAMObservacions, FAMbSeccio, FAMNIF,
                FAMDataNaixement, FAMQuota, FAMIDSec, FAMDataBaixa, FAMTipus,
                FAMSexe, FAMSociReferencia, FAMNewId, FAMNewIdRef,
                FAMbPagamentDomiciliat, FAMbRebutCobrat, FAMPagamentFinestreta
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            cursor.execute(query, (
                nuevo_famid,  # Usar el nuevo FAMID generado
                socio['FAMNom'],
                socio['FAMAdressa'],
                socio['FAMPoblacio'],
                socio['FAMCodPos'],
                socio['FAMTelefon'],
                socio['FAMMobil'],
                socio['FAMEmail'],
                socio['FAMDataAlta'],
                '',  # FAMCCC
                socio['FAMIBAN'],
                socio['FAMBIC'],
                0,  # FAMNSocis
                0,  # bBaixa (activo)
                '',  # FAMObservacions
                0,  # FAMbSeccio
                socio['FAMNIF'],
                None,  # FAMDataNaixement
                0,  # FAMQuota
                0,  # FAMIDSec
                None,  # FAMDataBaixa
                0,  # FAMTipus
                '',  # FAMSexe
                0,  # FAMSociReferencia
                0,  # FAMNewId
                0,  # FAMNewIdRef
                socio['FAMbPagamentDomiciliat'],
                0,  # FAMbRebutCobrat
                0   # FAMPagamentFinestreta
            ))
            
            self.conn.commit()
            self.stats['nuevos'] += 1
            print(f"  ➕ Nuevo socio: {socio['FAMNom'][:40]} (NIF: {socio['NIF_LIMPIO']}, ID: {nuevo_famid})")
            return True
            
        except Exception as e:
            print(f"  ❌ Error al insertar {socio['FAMNom']}: {e}")
            self.stats['errores'] += 1
            return False
    
    def actualizar_socio(self, socio, famid_bd):
        """Actualiza un socio existente en la base de datos."""
        try:
            cursor = self.conn.cursor()
            
            query = """
            UPDATE scazorla_sa.G_Socis SET
                FAMNom = ?,
                FAMAdressa = ?,
                FAMPoblacio = ?,
                FAMCodPos = ?,
                FAMTelefon = ?,
                FAMMobil = ?,
                FAMEmail = ?,
                FAMDataAlta = ?,
                FAMIBAN = ?,
                FAMBIC = ?,
                FAMNIF = ?,
                bBaixa = 0,
                FAMDataBaixa = NULL,
                FAMbPagamentDomiciliat = ?
            WHERE FAMID = ?
            """
            
            cursor.execute(query, (
                socio['FAMNom'],
                socio['FAMAdressa'],
                socio['FAMPoblacio'],
                socio['FAMCodPos'],
                socio['FAMTelefon'],
                socio['FAMMobil'],
                socio['FAMEmail'],
                socio['FAMDataAlta'],
                socio['FAMIBAN'],
                socio['FAMBIC'],
                socio['FAMNIF'],
                socio['FAMbPagamentDomiciliat'],
                famid_bd
            ))
            
            if cursor.rowcount > 0:
                self.conn.commit()
                self.stats['actualizados'] += 1
                print(f"  🔄 Actualizado: {socio['FAMNom'][:40]} (NIF: {socio['NIF_LIMPIO']})")
                return True
            else:
                self.stats['sin_cambios'] += 1
                return True
            
        except Exception as e:
            print(f"  ❌ Error al actualizar {socio['FAMNom']}: {e}")
            self.stats['errores'] += 1
            return False
    
    def marcar_bajas(self, nifs_excel):
        """
        Marca como baja los socios que NO están en el Excel.
        Ahora usa NIF en lugar de FAMID para identificar.
        
        Args:
            nifs_excel (set): Conjunto de NIFs de socios del Excel
        """
        print("\n🔍 Marcando como baja los socios que no están en el Excel...")
        
        try:
            cursor = self.conn.cursor()
            
            # Obtener socios activos en BD
            cursor.execute("""
                SELECT FAMID, FAMNom, FAMNIF
                FROM scazorla_sa.G_Socis 
                WHERE bBaixa = 0 OR bBaixa IS NULL
            """)
            
            socios_para_baja = []
            for row in cursor.fetchall():
                famid, nombre, famnif = row
                nif_limpio = self.limpiar_nif(famnif)
                
                # Solo procesar socios con NIF válido
                if nif_limpio and nif_limpio not in nifs_excel:
                    socios_para_baja.append((famid, nombre, nif_limpio))
            
            if not socios_para_baja:
                print("  ✅ No hay socios para marcar como baja")
                return
            
            print(f"  ⚠️  Se marcarán {len(socios_para_baja)} socios como baja:")
            
            # Marcar como baja
            fecha_baja = datetime.now()
            for famid, nombre, nif in socios_para_baja:
                cursor.execute("""
                    UPDATE scazorla_sa.G_Socis 
                    SET bBaixa = 1, FAMDataBaixa = ?
                    WHERE FAMID = ?
                """, (fecha_baja, famid))
                
                self.stats['marcados_baja'] += 1
                if self.stats['marcados_baja'] <= 10:
                    print(f"     ⚠️  {nombre[:40]} (NIF: {nif})")
            
            if len(socios_para_baja) > 10:
                print(f"     ... y {len(socios_para_baja) - 10} más")
            
            self.conn.commit()
            print(f"\n✅ Total de socios marcados como baja: {self.stats['marcados_baja']}")
            
        except Exception as e:
            print(f"❌ Error al marcar bajas: {e}")
            self.stats['errores'] += 1
    
    def sincronizar(self):
        """Ejecuta el proceso completo de sincronización."""
        print("\n" + "="*70)
        print("🔄 SINCRONIZACIÓN DE SOCIOS POR NIF/DNI")
        print("="*70)
        
        # 1. Crear backup
        backup_file = self.crear_backup()
        if backup_file:
            print(f"💾 Backup guardado: {backup_file}\n")
        
        # 2. Leer socios del Excel
        socios_excel = self.leer_excel()
        
        # 3. Obtener socios de la BD (indexados por NIF)
        socios_bd = self.obtener_socios_bd_por_nif()
        
        # 4. Crear set de NIFs del Excel para búsquedas rápidas
        nifs_excel = set(socio['NIF_LIMPIO'] for socio in socios_excel)
        
        # 5. Insertar o actualizar socios del Excel
        print(f"\n🔄 Procesando {len(socios_excel)} socios del Excel...")
        for i, socio in enumerate(socios_excel, 1):
            nif_limpio = socio['NIF_LIMPIO']
            
            if nif_limpio in socios_bd:
                # Socio existe: actualizar
                famid_bd = socios_bd[nif_limpio]['FAMID']
                self.actualizar_socio(socio, famid_bd)
            else:
                # Socio nuevo: insertar
                self.insertar_socio(socio)
            
            # Mostrar progreso cada 100 socios
            if i % 100 == 0:
                print(f"  ... procesados {i}/{len(socios_excel)}")
        
        # 6. Marcar como baja los socios que no están en el Excel
        self.marcar_bajas(nifs_excel)
        
        # 7. Obtener estadísticas finales
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM scazorla_sa.G_Socis")
        self.stats['total_bd_despues'] = cursor.fetchone()[0]
        
        # 8. Mostrar resumen
        self.mostrar_resumen()
    
    def mostrar_resumen(self):
        """Muestra un resumen de la sincronización."""
        print("\n" + "="*70)
        print("📊 RESUMEN DE LA SINCRONIZACIÓN")
        print("="*70)
        print(f"📂 Socios en Excel:                    {self.stats['total_excel']}")
        print(f"💾 Socios en BD (antes):               {self.stats['total_bd_antes']}")
        print(f"💾 Socios en BD (después):             {self.stats['total_bd_despues']}")
        print(f"-"*70)
        print(f"➕ Socios nuevos insertados:           {self.stats['nuevos']}")
        print(f"🔄 Socios actualizados:                {self.stats['actualizados']}")
        print(f"- Socios sin cambios:                  {self.stats['sin_cambios']}")
        print(f"⚠️  Socios marcados como baja:         {self.stats['marcados_baja']}")
        if self.stats['sin_nif'] > 0:
            print(f"⚠️  Socios omitidos (sin NIF):         {self.stats['sin_nif']}")
        if self.stats['nifs_duplicados'] > 0:
            print(f"⚠️  NIFs duplicados en Excel:          {self.stats['nifs_duplicados']}")
        print(f"❌ Errores:                            {self.stats['errores']}")
        print("="*70)
        
        if self.stats['errores'] == 0:
            print("✅ SINCRONIZACIÓN COMPLETADA EXITOSAMENTE")
        else:
            print(f"⚠️  SINCRONIZACIÓN COMPLETADA CON {self.stats['errores']} ERRORES")
        print("="*70 + "\n")
    
    def cerrar(self):
        """Cierra la conexión a la base de datos."""
        if self.conn:
            self.conn.close()
            print("🔒 Conexión a la base de datos cerrada")


def main():
    """Función principal."""
    # Configuración
    EXCEL_PATH = "Socis-2025.xlsx"
    ENV_PATH = "models/.env"
    
    print("="*70)
    print("🚀 SINCRONIZACIÓN DE SOCIOS POR NIF/DNI")
    print("="*70)
    print(f"📅 Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("="*70)
    
    # Verificar que el archivo Excel existe
    if not os.path.exists(EXCEL_PATH):
        print(f"\n❌ Error: No se encuentra el archivo {EXCEL_PATH}")
        print(f"   Coloca el archivo Excel en la carpeta del proyecto")
        sys.exit(1)
    
    # Verificar que el archivo .env existe
    if not os.path.exists(ENV_PATH):
        print(f"\n❌ Error: No se encuentra el archivo {ENV_PATH}")
        print(f"   Crea el archivo .env con las credenciales de la base de datos")
        sys.exit(1)
    
    # Pedir confirmación
    print("\n⚠️  ADVERTENCIA:")
    print("   • Este script usa el NIF como identificador único")
    print("   • Actualizará los socios existentes (por NIF)")
    print("   • Insertará nuevos socios")
    print("   • Marcará como BAJA los que NO estén en el Excel")
    print("   • Se creará un backup automático antes de empezar")
    
    respuesta = input("\n¿Deseas continuar? (escribe 'SI' para confirmar): ")
    
    if respuesta.upper() != 'SI':
        print("\n❌ Sincronización cancelada por el usuario")
        sys.exit(0)
    
    # Ejecutar sincronización
    try:
        sincronizador = SincronizadorSociosPorNIF(EXCEL_PATH, ENV_PATH)
        sincronizador.sincronizar()
        sincronizador.cerrar()
    except KeyboardInterrupt:
        print("\n\n⚠️  Sincronización interrumpida por el usuario")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()