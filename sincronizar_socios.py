#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Sincronización de Socios desde Excel a Base de Datos SQL Server

Este script:
1. Lee los socios activos del archivo Excel
2. Actualiza o inserta los socios en la base de datos
3. Marca como baja (bBaixa=1) los socios que NO aparecen en el Excel
4. Genera un reporte de los cambios realizados

Autor: Sistema de Gestión COJUB
Fecha: 2024
"""

import os
import sys
from datetime import datetime
from dotenv import load_dotenv
import pyodbc
import openpyxl

# Cargar variables de entorno
load_dotenv()

class SincronizadorSocios:
    def __init__(self, excel_path, env_path='models/.env'):
        """
        Inicializa el sincronizador.
        
        Args:
            excel_path (str): Ruta al archivo Excel con los socios activos
            env_path (str): Ruta al archivo .env con credenciales de BD
        """
        self.excel_path = excel_path
        self.conn = None
        self.stats = {
            'nuevos': 0,
            'actualizados': 0,
            'marcados_baja': 0,
            'errores': 0,
            'total_excel': 0,
            'total_bd_antes': 0,
            'total_bd_despues': 0
        }
        
        # Cargar variables de entorno desde la ruta especificada
        if os.path.exists(env_path):
            load_dotenv(env_path)
        
        self.conectar_bd()
    
    def conectar_bd(self):
        """Establece conexión con SQL Server."""
        try:
            conn_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={os.getenv('SQL_SERVER')};"
                f"DATABASE={os.getenv('SQL_DATABASE')};"
                f"UID={os.getenv('SQL_USER')};"
                f"PWD={os.getenv('SQL_PASSWORD')};"
            )
            self.conn = pyodbc.connect(conn_str)
            print("✅ Conexión a la base de datos establecida correctamente")
        except pyodbc.Error as ex:
            print(f"❌ Error al conectar a la base de datos: {ex}")
            sys.exit(1)
    
    def leer_excel(self):
        """
        Lee el archivo Excel y retorna una lista de diccionarios con los socios.
        
        Returns:
            list: Lista de diccionarios con datos de socios
        """
        print(f"\n📂 Leyendo archivo Excel: {self.excel_path}")
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb['Hoja1']
            
            socios_excel = []
            
            # Leer desde la fila 2 (saltando encabezados)
            for row in range(2, ws.max_row + 1):
                # Obtener valores
                codi = ws.cell(row, 1).value  # Codi
                nombre = ws.cell(row, 2).value  # Nombre
                nif = ws.cell(row, 3).value  # NIF
                direccion = ws.cell(row, 4).value  # Dirección
                cp = ws.cell(row, 5).value  # CP
                poblacion = ws.cell(row, 7).value  # Población
                telefono = ws.cell(row, 9).value  # Teléfono
                movil = ws.cell(row, 10).value  # Móvil
                email = ws.cell(row, 11).value  # Email
                forma_pago = ws.cell(row, 12).value  # Forma de pagament
                iban = ws.cell(row, 13).value  # IBAN
                bic = ws.cell(row, 14).value  # BIC
                fecha_alta = ws.cell(row, 15).value  # Fecha Alta
                
                # Validar que al menos tenga código
                if not codi:
                    continue
                
                # Convertir Codi a string
                codi = str(codi).strip()
                
                # Determinar si el pago está domiciliado
                pago_domiciliado = False
                if forma_pago and isinstance(forma_pago, str):
                    pago_domiciliado = 'domiciliat' in forma_pago.lower() or '3' in forma_pago
                
                # Convertir fecha de Excel a datetime si es numérico
                fecha_alta_dt = None
                if fecha_alta:
                    if isinstance(fecha_alta, (int, float)):
                        # Excel guarda fechas como números (días desde 1900-01-01)
                        try:
                            from datetime import timedelta
                            fecha_alta_dt = datetime(1899, 12, 30) + timedelta(days=int(fecha_alta))
                        except:
                            fecha_alta_dt = None
                    elif isinstance(fecha_alta, datetime):
                        fecha_alta_dt = fecha_alta
                
                # Convertir CP a string
                cp_str = str(cp) if cp else ""
                if cp_str and '.' in cp_str:
                    cp_str = cp_str.split('.')[0]  # Quitar decimales si los hay
                
                socio = {
                    'FAMID': codi,
                    'FAMNom': str(nombre).strip() if nombre else "",
                    'FAMNIF': str(nif).strip() if nif else "",
                    'FAMAdressa': str(direccion).strip() if direccion else "",
                    'FAMPoblacio': str(poblacion).strip() if poblacion else "",
                    'FAMCodPos': cp_str.strip(),
                    'FAMTelefon': str(telefono).strip() if telefono else "",
                    'FAMMobil': str(movil).strip() if movil else "",
                    'FAMEmail': str(email).strip() if email else "",
                    'FAMIBAN': str(iban).strip() if iban else "",
                    'FAMBIC': str(bic).strip() if bic else "",
                    'FAMDataAlta': fecha_alta_dt,
                    'FAMbPagamentDomiciliat': pago_domiciliado,
                    'bBaixa': False  # Los del Excel están activos
                }
                
                socios_excel.append(socio)
            
            self.stats['total_excel'] = len(socios_excel)
            print(f"✅ Se leyeron {len(socios_excel)} socios del Excel")
            
            return socios_excel
            
        except Exception as e:
            print(f"❌ Error al leer el Excel: {e}")
            sys.exit(1)
    
    def obtener_socios_bd(self):
        """
        Obtiene todos los IDs de socios de la base de datos.
        
        Returns:
            set: Conjunto con los IDs de socios en la BD
        """
        cursor = self.conn.cursor()
        cursor.execute("SELECT FAMID FROM scazorla_sa.G_Socis")
        socios_bd = set(row[0] for row in cursor.fetchall())
        self.stats['total_bd_antes'] = len(socios_bd)
        print(f"📊 Total de socios en BD antes de sincronizar: {len(socios_bd)}")
        return socios_bd
    
    def socio_existe(self, famid):
        """Verifica si un socio existe en la BD."""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM scazorla_sa.G_Socis WHERE FAMID = ?", famid)
        return cursor.fetchone()[0] > 0
    
    def insertar_socio(self, socio):
        """Inserta un nuevo socio en la base de datos."""
        try:
            cursor = self.conn.cursor()
            
            query = """
            INSERT INTO scazorla_sa.G_Socis (
                FAMID, FAMNom, FAMAdressa, FAMPoblacio, FAMCodPos, FAMTelefon,
                FAMMobil, FAMEmail, FAMDataAlta, FAMCCC, FAMIBAN, FAMBIC,
                FAMNSocis, bBaixa, FAMObservacions, FAMbSeccio, FAMNIF,
                FAMDataNaixement, FAMQuota, FAMIDSec, FAMDataBaixa, FAMTipus,
                FAMSexe, FAMSociReferencia, FAMNewId, FAMNewIdRef,
                FAMbPagamentDomiciliat, FAMbRebutCobrat, FAMPagamentFinestreta
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            cursor.execute(query, (
                socio['FAMID'],
                socio['FAMNom'],
                socio['FAMAdressa'],
                socio['FAMPoblacio'],
                socio['FAMCodPos'],
                socio['FAMTelefon'],
                socio['FAMMobil'],
                socio['FAMEmail'],
                socio['FAMDataAlta'],
                '',  # FAMCCC
                socio['FAMIBAN'],
                socio['FAMBIC'],
                0,  # FAMNSocis
                0,  # bBaixa (activo)
                '',  # FAMObservacions
                0,  # FAMbSeccio
                socio['FAMNIF'],
                None,  # FAMDataNaixement
                0.0,  # FAMQuota
                '',  # FAMIDSec
                None,  # FAMDataBaixa
                '',  # FAMTipus
                '',  # FAMSexe
                '',  # FAMSociReferencia
                '',  # FAMNewId
                '',  # FAMNewIdRef
                socio['FAMbPagamentDomiciliat'],
                0,  # FAMbRebutCobrat
                0   # FAMPagamentFinestreta
            ))
            
            self.conn.commit()
            self.stats['nuevos'] += 1
            print(f"  ➕ Nuevo socio insertado: {socio['FAMID']} - {socio['FAMNom']}")
            return True
            
        except Exception as e:
            print(f"  ❌ Error al insertar socio {socio['FAMID']}: {e}")
            self.stats['errores'] += 1
            return False
    
    def actualizar_socio(self, socio):
        """Actualiza un socio existente en la base de datos."""
        try:
            cursor = self.conn.cursor()
            
            query = """
            UPDATE scazorla_sa.G_Socis SET
                FAMNom = ?,
                FAMAdressa = ?,
                FAMPoblacio = ?,
                FAMCodPos = ?,
                FAMTelefon = ?,
                FAMMobil = ?,
                FAMEmail = ?,
                FAMDataAlta = ?,
                FAMIBAN = ?,
                FAMBIC = ?,
                FAMNIF = ?,
                bBaixa = 0,
                FAMDataBaixa = NULL,
                FAMbPagamentDomiciliat = ?
            WHERE FAMID = ?
            """
            
            cursor.execute(query, (
                socio['FAMNom'],
                socio['FAMAdressa'],
                socio['FAMPoblacio'],
                socio['FAMCodPos'],
                socio['FAMTelefon'],
                socio['FAMMobil'],
                socio['FAMEmail'],
                socio['FAMDataAlta'],
                socio['FAMIBAN'],
                socio['FAMBIC'],
                socio['FAMNIF'],
                socio['FAMbPagamentDomiciliat'],
                socio['FAMID']
            ))
            
            self.conn.commit()
            self.stats['actualizados'] += 1
            print(f"  🔄 Socio actualizado: {socio['FAMID']} - {socio['FAMNom']}")
            return True
            
        except Exception as e:
            print(f"  ❌ Error al actualizar socio {socio['FAMID']}: {e}")
            self.stats['errores'] += 1
            return False
    
    def marcar_bajas(self, ids_excel):
        """
        Marca como baja los socios que NO están en el Excel.
        
        Args:
            ids_excel (set): Conjunto de IDs de socios del Excel
        """
        print("\n🔍 Marcando como baja los socios que no están en el Excel...")
        
        try:
            cursor = self.conn.cursor()
            
            # Obtener socios activos en BD que no están en Excel
            cursor.execute("""
                SELECT FAMID, FAMNom 
                FROM scazorla_sa.G_Socis 
                WHERE bBaixa = 0
            """)
            
            socios_para_baja = []
            for row in cursor.fetchall():
                if row[0] not in ids_excel:
                    socios_para_baja.append((row[0], row[1]))
            
            if not socios_para_baja:
                print("  ✅ No hay socios para marcar como baja")
                return
            
            # Marcar como baja
            fecha_baja = datetime.now()
            for famid, nombre in socios_para_baja:
                cursor.execute("""
                    UPDATE scazorla_sa.G_Socis 
                    SET bBaixa = 1, FAMDataBaixa = ?
                    WHERE FAMID = ?
                """, (fecha_baja, famid))
                
                self.stats['marcados_baja'] += 1
                print(f"  ⚠️ Socio marcado como baja: {famid} - {nombre}")
            
            self.conn.commit()
            print(f"\n✅ Total de socios marcados como baja: {self.stats['marcados_baja']}")
            
        except Exception as e:
            print(f"❌ Error al marcar bajas: {e}")
            self.stats['errores'] += 1
    
    def sincronizar(self):
        """Ejecuta el proceso completo de sincronización."""
        print("\n" + "="*70)
        print("🔄 INICIANDO SINCRONIZACIÓN DE SOCIOS")
        print("="*70)
        
        # 1. Leer socios del Excel
        socios_excel = self.leer_excel()
        
        # 2. Obtener socios de la BD
        socios_bd = self.obtener_socios_bd()
        
        # 3. Crear set de IDs del Excel para búsquedas rápidas
        ids_excel = set(socio['FAMID'] for socio in socios_excel)
        
        # 4. Insertar o actualizar socios del Excel
        print(f"\n🔄 Procesando {len(socios_excel)} socios del Excel...")
        for socio in socios_excel:
            if self.socio_existe(socio['FAMID']):
                self.actualizar_socio(socio)
            else:
                self.insertar_socio(socio)
        
        # 5. Marcar como baja los socios que no están en el Excel
        self.marcar_bajas(ids_excel)
        
        # 6. Obtener estadísticas finales
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM scazorla_sa.G_Socis")
        self.stats['total_bd_despues'] = cursor.fetchone()[0]
        
        # 7. Mostrar resumen
        self.mostrar_resumen()
    
    def mostrar_resumen(self):
        """Muestra un resumen de la sincronización."""
        print("\n" + "="*70)
        print("📊 RESUMEN DE LA SINCRONIZACIÓN")
        print("="*70)
        print(f"📂 Socios en Excel:                    {self.stats['total_excel']}")
        print(f"💾 Socios en BD (antes):               {self.stats['total_bd_antes']}")
        print(f"💾 Socios en BD (después):             {self.stats['total_bd_despues']}")
        print(f"-"*70)
        print(f"➕ Socios nuevos insertados:           {self.stats['nuevos']}")
        print(f"🔄 Socios actualizados:                {self.stats['actualizados']}")
        print(f"⚠️  Socios marcados como baja:         {self.stats['marcados_baja']}")
        print(f"❌ Errores:                            {self.stats['errores']}")
        print("="*70)
        
        if self.stats['errores'] == 0:
            print("✅ SINCRONIZACIÓN COMPLETADA EXITOSAMENTE")
        else:
            print(f"⚠️  SINCRONIZACIÓN COMPLETADA CON {self.stats['errores']} ERRORES")
        print("="*70 + "\n")
    
    def cerrar(self):
        """Cierra la conexión a la base de datos."""
        if self.conn:
            self.conn.close()
            print("🔒 Conexión a la base de datos cerrada")


def main():
    """Función principal."""
    # Configuración
    EXCEL_PATH = "Socis-2025.xlsx"
    ENV_PATH = "models/.env"
    
    # Verificar que el archivo Excel existe
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Error: No se encuentra el archivo {EXCEL_PATH}")
        print(f"   Coloca el archivo Excel en la carpeta del proyecto")
        sys.exit(1)
    
    # Verificar que el archivo .env existe
    if not os.path.exists(ENV_PATH):
        print(f"❌ Error: No se encuentra el archivo {ENV_PATH}")
        print(f"   Crea el archivo .env con las credenciales de la base de datos")
        sys.exit(1)
    
    # Ejecutar sincronización
    try:
        sincronizador = SincronizadorSocios(EXCEL_PATH, ENV_PATH)
        sincronizador.sincronizar()
        sincronizador.cerrar()
    except KeyboardInterrupt:
        print("\n\n⚠️  Sincronización interrumpida por el usuario")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()